import openpyxl

wb = openpyxl.load_workbook("1.xlsx")
ws = wb.active


def combine_strings(text: str) -> str:
    result = ""
    lines = text.split('、')
    used_lines = set()  # 用于存储已经使用过的字符串
    for i, line in enumerate(lines):
        if line in used_lines:
            continue
        if i == 0:
            result += line
            used_lines.add(line)
        elif len(result) + len(line) + 1 <= 11 and line not in used_lines:
            result += '、' + line
            used_lines.add(line)
        else:
            break
    return result


def combine_two_strings(text: str) -> str:
    result = ""
    lines = text.split("、")
    used_lines = set()
    for line in lines:
        if len(used_lines) == 2:
            break
        if line in used_lines or len(line) > 5:
            continue
        if result and len(result) + len(line) + 1 > 6:
            result += "、\n"
        else:
            result += "、"
        result += line
        used_lines.add(line)
    return result.lstrip("、\n").rstrip("、")


def translate(word):
    for row in ws.iter_rows():
        # 遍历当前行的所有单元格
        for cell in row:
            # 检查单元格的值是否等于要查找的值
            if cell.value == word:
                # 获取该单元格右边的单元格
                right_cell = cell.offset(row=0, column=1)
                # wb.close()
                # 输出查找结果
                return right_cell.value


def check(word):
    # 遍历第一列，查找指定单词
    for cell in ws['A']:
        if cell.value == word:
            return True
    else:
        return False
